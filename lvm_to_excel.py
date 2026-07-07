"""
lvm_to_excel.py
---------------
Procesa archivos LVM (formato POS Aclas) y genera un reporte Excel en español.

Uso:
    python lvm_to_excel.py <archivo_entrada.txt> [archivo_salida.xlsx]

Ejemplo:
    python lvm_to_excel.py LVM06_1.txt
    python lvm_to_excel.py LVM06_1.txt Reporte_Junio.xlsx

Requisitos:
    pip install openpyxl
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color palette ────────────────────────────────────────────────
HDR_BG   = '1F3864'
HDR_FG   = 'FFFFFF'
ALT_BG   = 'E8EFF8'
SUMM_BG  = 'DCE6F1'
TOTAL_BG = 'BDD7EE'
TITLE_FG = '1F3864'

MONEY_FMT = '#,##0.00'
INT_FMT   = '#,##0'

_thin  = Side(style='thin',   color='B8CCE4')
_thick = Side(style='medium', color='1F3864')
B_THIN  = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
B_THICK = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thick)


# ── Style helpers ────────────────────────────────────────────────
def style_header(cell):
    cell.font      = Font(name='Arial', bold=True, color=HDR_FG, size=10)
    cell.fill      = PatternFill('solid', start_color=HDR_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = B_THIN

def style_data(cell, row_idx):
    bg = ALT_BG if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = B_THIN
    cell.font      = Font(name='Arial', size=9)

def style_money(cell, row_idx):
    bg = ALT_BG if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill         = PatternFill('solid', start_color=bg)
    cell.alignment    = Alignment(horizontal='right', vertical='center')
    cell.border       = B_THIN
    cell.number_format = MONEY_FMT
    cell.font         = Font(name='Arial', size=9)

def style_total(cell, fmt=MONEY_FMT):
    cell.font         = Font(name='Arial', bold=True, size=9, color=TITLE_FG)
    cell.fill         = PatternFill('solid', start_color=TOTAL_BG)
    cell.alignment    = Alignment(horizontal='right', vertical='center')
    cell.border       = B_THICK
    cell.number_format = fmt

def style_total_empty(cell):
    cell.fill   = PatternFill('solid', start_color=TOTAL_BG)
    cell.border = B_THICK
    cell.font   = Font(name='Arial', bold=True, size=9)

def make_title(ws, text, col_span):
    ws.merge_cells(f'A1:{get_column_letter(col_span)}1')
    cell = ws['A1']
    cell.value     = text
    cell.font      = Font(name='Arial', bold=True, size=14, color=TITLE_FG)
    cell.fill      = PatternFill('solid', start_color=SUMM_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28


# ── Parsing ──────────────────────────────────────────────────────
def fmt_date(raw):
    return f"{raw[:4]}-{raw[4:6]}-{raw[6:]}" if len(raw) == 8 else raw

def fmt_time(raw):
    return f"{raw[:2]}:{raw[2:4]}:{raw[4:]}" if len(raw) == 6 else raw

def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def safe_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0

DOC_TYPE = {'1': 'Venta', '2': 'Devolución'}
PAYMENT  = {'': 'Efectivo', '2': 'Tarjeta'}

def parse_lvm(filepath):
    summary      = None
    batches      = []
    transactions = []

    with open(filepath, encoding='utf-8') as f:
        for line in f:
            parts = line.strip().split('||')
            rtype = parts[0]

            if rtype == '3':
                summary = parts

            elif rtype == '1' and parts[3]:   # real batch (has invoice count)
                batches.append({
                    'store_id':      parts[2],
                    'invoice_count': safe_int(parts[3]),
                    'gross_amount':  safe_float(parts[4]),
                    'tax_amount':    safe_float(parts[5]),
                    'net_amount':    safe_float(parts[7]),
                    'taxable_gross': safe_float(parts[11]),
                    'taxable_tax':   safe_float(parts[12]),
                    'exempt_gross':  safe_float(parts[14]),
                    'first_invoice': parts[24],
                    'last_invoice':  parts[25],
                    'batch_number':  parts[26],
                    'device':        parts[27],
                    'version':       parts[28],
                })

            elif rtype == '2':
                transactions.append({
                    'cashier':        parts[1],
                    'terminal':       parts[2],
                    'invoice_number': parts[4],
                    'date':           fmt_date(parts[5]),
                    'time':           fmt_time(parts[6]),
                    'doc_type':       DOC_TYPE.get(parts[7], parts[7]),
                    'payment':        PAYMENT.get(parts[8], parts[8]),
                    'fiscal_receipt': parts[10],
                    'gross_amount':   safe_float(parts[12]),
                    'tax_amount':     safe_float(parts[13]),
                    'net_amount':     safe_float(parts[15]),
                    'item_count':     safe_int(parts[21]),
                    'taxable_amount': safe_float(parts[30]),
                })

    return summary, batches, transactions


# ── Sheet builders ───────────────────────────────────────────────
def build_transactions_sheet(ws, summary, batches, transactions):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'

    file_id = os.path.splitext(os.path.basename(summary[0] if summary else 'LVM'))[0]
    make_title(ws, f'Reporte de Transacciones', 13)

    ws.merge_cells('A2:M2')
    subtitle = ws['A2']
    subtitle.value = (
        f"Tienda: {batches[0]['store_id']}   |   "
        f"Dispositivo: {batches[0]['device']} {batches[0]['version']}   |   "
        f"Total Facturas: {summary[2]}"
    )
    subtitle.font      = Font(name='Arial', size=9, italic=True, color='595959')
    subtitle.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    headers = [
        'No. Factura', 'Fecha', 'Hora', 'Tipo Doc.', 'Pago',
        'No. Comprobante Fiscal', 'Monto Bruto (RD$)', 'ITBIS (RD$)',
        'Monto Neto (RD$)', 'Cant. Artículos', 'Monto Gravable (RD$)',
        'ID Cajero', 'Terminal'
    ]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=col, value=h))
    ws.row_dimensions[3].height = 30

    for i, txn in enumerate(transactions):
        row  = i + 4
        vals = [
            txn['invoice_number'], txn['date'], txn['time'],
            txn['doc_type'], txn['payment'], txn['fiscal_receipt'],
            txn['gross_amount'], txn['tax_amount'], txn['net_amount'],
            txn['item_count'], txn['taxable_amount'],
            txn['cashier'], txn['terminal'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (7, 8, 9, 11):
                style_money(cell, i)
            elif col == 10:
                style_data(cell, i)
                cell.number_format = INT_FMT
            else:
                style_data(cell, i)

    # Totals row
    trow = len(transactions) + 4
    lbl = ws.cell(row=trow, column=1, value='TOTAL')
    lbl.font      = Font(name='Arial', bold=True, size=9, color=TITLE_FG)
    lbl.fill      = PatternFill('solid', start_color=TOTAL_BG)
    lbl.alignment = Alignment(horizontal='center')
    lbl.border    = B_THICK

    total_cols = {7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K'}
    for col in range(2, 14):
        cell = ws.cell(row=trow, column=col)
        if col in total_cols:
            letter = total_cols[col]
            cell.value = f'=SUM({letter}4:{letter}{trow-1})'
            style_total(cell, INT_FMT if col == 10 else MONEY_FMT)
        else:
            style_total_empty(cell)

    col_widths = [22, 12, 10, 11, 11, 24, 15, 13, 13, 14, 16, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_batches_sheet(ws, batches):
    ws.sheet_view.showGridLines = False
    make_title(ws, 'Resumen de Lotes', 11)

    headers = [
        'No. Lote', 'ID Tienda', 'Dispositivo', 'Cant. Facturas',
        'Primera Factura', 'Última Factura',
        'Monto Bruto (RD$)', 'ITBIS (RD$)', 'Monto Neto (RD$)',
        'Gravable Bruto (RD$)', 'Exento Bruto (RD$)'
    ]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=2, column=col, value=h))
    ws.row_dimensions[2].height = 30

    for i, b in enumerate(batches):
        row  = i + 3
        vals = [
            b['batch_number'], b['store_id'],
            f"{b['device']} {b['version']}",
            b['invoice_count'], b['first_invoice'], b['last_invoice'],
            b['gross_amount'], b['tax_amount'], b['net_amount'],
            b['taxable_gross'], b['exempt_gross'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (7, 8, 9, 10, 11):
                style_money(cell, i)
            else:
                style_data(cell, i)

    trow = len(batches) + 3
    lbl = ws.cell(row=trow, column=1, value='TOTAL')
    lbl.font      = Font(name='Arial', bold=True, size=9, color=TITLE_FG)
    lbl.fill      = PatternFill('solid', start_color=TOTAL_BG)
    lbl.alignment = Alignment(horizontal='center')
    lbl.border    = B_THICK

    total_cols = {4: 'D', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K'}
    for col in range(2, 12):
        cell = ws.cell(row=trow, column=col)
        if col in total_cols:
            letter = total_cols[col]
            cell.value = f'=SUM({letter}3:{letter}{trow-1})'
            style_total(cell, INT_FMT if col == 4 else MONEY_FMT)
        else:
            style_total_empty(cell)

    col_widths = [10, 12, 16, 14, 24, 24, 15, 14, 14, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_summary_sheet(ws, summary):
    ws.sheet_view.showGridLines = False
    make_title(ws, 'Resumen General', 3)

    rows = [
        ('Total de Transacciones',           safe_int(summary[2]),    INT_FMT),
        ('Monto Bruto Total (RD$)',           safe_float(summary[3]),  MONEY_FMT),
        ('ITBIS Total (RD$)',                 safe_float(summary[4]),  MONEY_FMT),
        ('Monto Neto Total (RD$)',            safe_float(summary[6]),  MONEY_FMT),
        ('Base Imponible (RD$)',              safe_float(summary[10]), MONEY_FMT),
        ('ITBIS sobre Base Imponible (RD$)',  safe_float(summary[11]), MONEY_FMT),
        ('Monto Exento (RD$)',                safe_float(summary[12]), MONEY_FMT),
        ('ITBIS sobre Exento (RD$)',          safe_float(summary[13]), MONEY_FMT),
    ]

    for r, (label, val, fmt) in enumerate(rows, 2):
        ws.row_dimensions[r].height = 18
        lc = ws.cell(row=r, column=1, value=label)
        lc.font      = Font(name='Arial', bold=True, size=10, color=TITLE_FG)
        lc.fill      = PatternFill('solid', start_color=SUMM_BG)
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        lc.border    = B_THIN
        vc = ws.cell(row=r, column=2, value=val)
        vc.font          = Font(name='Arial', size=10)
        vc.alignment     = Alignment(horizontal='right', vertical='center')
        vc.border        = B_THIN
        vc.number_format = fmt
        ws.cell(row=r, column=3).border = B_THIN

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 5


# ── Main ─────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Uso: python lvm_to_excel.py <archivo.txt> [salida.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]

    if not os.path.exists(input_file):
        print(f"Error: no se encontró el archivo '{input_file}'")
        sys.exit(1)

    # Default output name: same name as input with .xlsx extension
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base}_Reporte.xlsx"

    print(f"Procesando: {input_file}")
    summary, batches, transactions = parse_lvm(input_file)

    if not summary:
        print("Error: no se encontró el registro de resumen (tipo 3) en el archivo.")
        sys.exit(1)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Transacciones'
    build_transactions_sheet(ws1, summary, batches, transactions)

    ws2 = wb.create_sheet('Resumen de Lotes')
    build_batches_sheet(ws2, batches)

    ws3 = wb.create_sheet('Resumen General')
    build_summary_sheet(ws3, summary)

    wb.save(output_file)
    print(f"Reporte guardado: {output_file}")
    print(f"  → {len(transactions)} transacciones | {len(batches)} lotes")


if __name__ == '__main__':
    main()
